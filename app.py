"""
Placement Management System - Flask Application
Complete web application for managing placements with Admin and Teacher roles
"""
import os
import secrets
import atexit
import subprocess
import time
import threading
from queue import Queue
from datetime import datetime
from functools import wraps
from pathlib import Path
import io

# Load environment variables from .env file if it exists
from dotenv import load_dotenv
env_path = os.path.join(os.path.dirname(__file__), '.env')
if os.path.exists(env_path):
    load_dotenv(env_path)

from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
from flask_session import Session
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash

import sys
from pathlib import Path

# Add src/python to sys.path for imports
project_root = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(project_root, 'src', 'python'))

from models import db, User, Teacher, Class, Student, UserSession
from voice_generator import create_placement_notification_voice
from whatsapp_automation import (
    send_placement_whatsapp,
    initialize_whatsapp,
    send_placement_call_with_twilio,
    get_placement_notifications,
)
from twilio_calling import make_placement_call





# Configuration
UPLOAD_FOLDER = 'static/uploads'
AUDIO_FOLDER = 'static/audio'
ALLOWED_EXTENSIONS = {'pdf', 'doc', 'docx', 'jpg', 'jpeg', 'png', 'xlsx', 'xls', 'csv'}
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# Initialize Flask app
app = Flask(__name__, instance_relative_config=True)

# Ensure instance folder exists
try:
    os.makedirs(app.instance_path, exist_ok=True)
except OSError:
    pass

# Use persistent secret key from environment or default to a generated one
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY') or secrets.token_hex(32)

# Session configuration for multiple concurrent users
app.config['SESSION_TYPE'] = 'filesystem'
app.config['PERMANENT_SESSION_LIFETIME'] = 86400  # 24 hours
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SECURE'] = False  # Set to True if using HTTPS
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['SESSION_COOKIE_NAME'] = 'placement_session'
app.config['SESSION_REFRESH_EACH_REQUEST'] = False

app.config['SQLALCHEMY_DATABASE_URI'] = os.getenv('SQLALCHEMY_DATABASE_URI') or f'sqlite:///{os.path.join(app.instance_path, "database.db")}'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_FILE_SIZE

# Create upload folders
Path(UPLOAD_FOLDER).mkdir(parents=True, exist_ok=True)
Path(AUDIO_FOLDER).mkdir(parents=True, exist_ok=True)

# Configure session storage directory
app.config['SESSION_FILE_DIR'] = os.path.join(app.instance_path, 'flask_session')
Path(app.config['SESSION_FILE_DIR']).mkdir(parents=True, exist_ok=True)

# Initialize database
db.init_app(app)

# Initialize session for multiple concurrent users
session_obj = Session()
session_obj.init_app(app)


# ==================== Database Initialization ====================

def clear_old_sessions():
    """Clear all session files on application startup to force re-login"""
    session_dir = app.config.get('SESSION_FILE_DIR')
    if session_dir and os.path.exists(session_dir):
        try:
            for filename in os.listdir(session_dir):
                file_path = os.path.join(session_dir, filename)
                if os.path.isfile(file_path):
                    os.remove(file_path)
            print("[SESSION] [*] Old sessions cleared - Users must login again")
        except Exception as e:
            print("[SESSION] [*][*]  Error clearing sessions: {}".format(str(e)))


def init_db():
    """Initialize the database with all tables and default admin"""
    with app.app_context():
        db.create_all()
        create_default_admin()


# ==================== Helper Functions ====================

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def admin_login_required(f):
    """Decorator to require admin login - checks admin_id session variable only"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_id' not in session or not session.get('admin_id'):
            flash('Admin login required.', 'warning')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function


def teacher_login_required(f):
    """Decorator to require teacher login - checks teacher_id session variable only"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'teacher_id' not in session or not session.get('teacher_id'):
            flash('Teacher login required.', 'warning')
            return redirect(url_for('teacher_login'))
        return f(*args, **kwargs)
    return decorated_function


def login_required(f):
    """Generic login decorator - checks for either admin or teacher session"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_id' not in session and 'teacher_id' not in session:
            flash('Please login first.', 'warning')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function


def admin_required(f):
    """Decorator to require admin role - uses admin_id session variable only"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'admin_id' not in session or not session.get('admin_id'):
            flash('Admin access required.', 'danger')
            return redirect(url_for('admin_login'))
        admin_user = db.session.get(User, session.get('admin_id'))
        if not admin_user or admin_user.role != 'admin':
            session.pop('admin_id', None)
            flash('Admin account no longer valid.', 'danger')
            return redirect(url_for('admin_login'))
        return f(*args, **kwargs)
    return decorated_function


def teacher_required(f):
    """Decorator to require teacher role - uses teacher_id session variable only"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'teacher_id' not in session or not session.get('teacher_id'):
            flash('Teacher access required.', 'danger')
            return redirect(url_for('teacher_login'))
        teacher_user = db.session.get(User, session.get('teacher_id'))
        if not teacher_user or teacher_user.role != 'teacher':
            session.pop('teacher_id', None)
            flash('Teacher account no longer valid.', 'danger')
            return redirect(url_for('teacher_login'))
        return f(*args, **kwargs)
    return decorated_function


def create_default_admin():
    """Create default admin user if not exists"""
    admin = User.query.filter_by(username='admin', role='admin').first()
    if not admin:
        admin = User(username='admin', role='admin')
        admin.set_password('admin')
        db.session.add(admin)
        db.session.commit()
        print("[INFO] Default admin user created: admin/admin")


def trigger_placement_notification(student):
    """
    Trigger placement notification when company name and package are filled
    Sends: WhatsApp message + Voice message + Phone call (if Twilio configured)
    Respects student's send_whatsapp preference
    """
    try:
        if not student.company_name or not student.package:
            return False
        
        if student.notification_sent:
            return False  # Already sent
        
        # Check if student wants WhatsApp notifications
        if not student.send_whatsapp:
            print(f"[INFO] WhatsApp notifications disabled for student: {student.full_name}")
            return False
        
        # Enqueue whatsapp + call tasks for background processing
        enqueue_whatsapp_task(student.id)
        enqueue_call_task(student.id)
        whatsapp_success = None
        call_result = None
        call_success = None

        # Update notification status (whatsapp + call metadata)
        try:
            # We mark tasks queued; workers will update to actual status once processed
            student.last_whatsapp_status = 'queued'
            student.last_call_status = 'queued'
            # Keep using notification_sent flag for general notifications
            if whatsapp_success:
                student.notification_sent = True
                student.notification_sent_at = datetime.utcnow()
            db.session.commit()
            return whatsapp_success or call_success
        except Exception as e:
            db.session.rollback()
            print(f"Error committing notification status to DB: {e}")
            return whatsapp_success or call_success
        
    except Exception as e:
        print(f"Error triggering placement notification: {e}")
        return False


# ==================== Application Context ====================

@app.before_request
def before_request():
    """Initialize database tables and create default admin"""
    db.create_all()
    create_default_admin()


# ==================== Authentication Routes ====================

# NOTE: /login alias intentionally redirects to the landing page. Kept for backwards compatibility.
@app.route('/login')
def login():
    return redirect(url_for('index'))


@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    """Admin login - stores only admin_id in session, preserves teacher_id"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        user = User.query.filter_by(username=username, role='admin').first()
        
        if user and user.check_password(password) and user.is_active:
            # Store ONLY admin_id - never clear teacher_id
            session.permanent = True
            session['admin_id'] = user.id
            session['role'] = 'admin'
            session['admin_username'] = user.username
            
            flash(f'Welcome Admin {username}!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid admin credentials.', 'danger')
    
    return render_template('admin_login.html')


@app.route('/teacher/login', methods=['GET', 'POST'])
def teacher_login():
    """Teacher login - stores only teacher_id in session, preserves admin_id"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        
        user = User.query.filter_by(username=username, role='teacher').first()
        
        if user and user.check_password(password) and user.is_active:
            # Store ONLY teacher_id - never clear admin_id
            session.permanent = True
            session['teacher_id'] = user.id
            session['role'] = 'teacher'
            session['teacher_username'] = user.username
            
            flash(f'Welcome Teacher {username}!', 'success')
            return redirect(url_for('teacher_dashboard'))
        else:
            flash('Invalid teacher credentials.', 'danger')
    
    return render_template('teacher_login.html')


@app.route('/admin/logout')
@admin_required
def admin_logout():
    """Logout from admin session ONLY - does NOT clear teacher session"""
    username = session.pop('admin_id', None)
    session.pop('admin_username', None)
    # If teacher still logged in, preserve role, otherwise clear it
    if session.get('teacher_id'):
        session['role'] = 'teacher'
    else:
        session.pop('role', None)
    flash('Admin logged out successfully.', 'info')
    # Redirect to home, which will show teacher dashboard if teacher is logged in
    return redirect(url_for('index'))


@app.route('/teacher/logout')
@teacher_required
def teacher_logout():
    """Logout from teacher session ONLY - does NOT clear admin session"""
    username = session.pop('teacher_id', None)
    session.pop('teacher_username', None)
    # If admin still logged in, preserve role, otherwise clear it
    if session.get('admin_id'):
        session['role'] = 'admin'
    else:
        session.pop('role', None)
    flash('Teacher logged out successfully.', 'info')
    # Redirect to home, which will show admin dashboard if admin is logged in
    return redirect(url_for('index'))


@app.route('/logout')
def logout():
    """Logout from both sessions"""
    session.pop('admin_id', None)
    session.pop('admin_username', None)
    session.pop('teacher_id', None)
    session.pop('teacher_username', None)
    session.pop('role', None)
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))



# ==================== Dashboard Routes ====================

@app.route('/')
def index():
    """Home page - redirect to appropriate dashboard based on session"""
    # Check admin session first
    if 'admin_id' in session and session.get('admin_id'):
        return redirect(url_for('admin_dashboard'))
    # Check teacher session
    if 'teacher_id' in session and session.get('teacher_id'):
        return redirect(url_for('teacher_dashboard'))
    # Neither logged in, show the new landing index page
    return render_template('index.html')


@app.route('/dashboard')
def dashboard():
    """Generic dashboard - redirect to appropriate dashboard"""
    if 'admin_id' in session and session.get('admin_id'):
        return redirect(url_for('admin_dashboard'))
    if 'teacher_id' in session and session.get('teacher_id'):
        return redirect(url_for('teacher_dashboard'))
    return redirect(url_for('index'))


@app.route('/admin/dashboard')
@admin_required
def admin_dashboard():
    """Admin dashboard - checks admin_id session variable ONLY"""
    # Get admin user from session
    admin_user = db.session.get(User, session.get('admin_id'))
    
    # Get search queries
    search_teacher = request.args.get('search_teacher', '').strip()
    search_class = request.args.get('search_class', '').strip()
    search_student = request.args.get('search_student', '').strip()
    
    teachers = Teacher.query.all()
    classes = Class.query.all()
    students = Student.query.all()
    
    # Filter teachers
    if search_teacher:
        teachers = [t for t in teachers if 
                   search_teacher.lower() in t.user.username.lower() or
                   search_teacher.lower() in t.full_name.lower() or
                   search_teacher.lower() in (t.email or '').lower()]
    
    # Filter classes
    if search_class:
        classes = [c for c in classes if 
                  search_class.lower() in c.course.lower() or
                  search_class.lower() in c.specialisation.lower()]
    
    # Filter students
    if search_student:
        students = [s for s in students if 
                   search_student.lower() in s.usn.lower() or
                   search_student.lower() in s.full_name.lower()]
    
    context = {
        'teachers_count': len(teachers),
        'classes_count': len(classes),
        'students_count': len(students),
        'placed_count': len([s for s in students if s.company_name]),
        'teachers': teachers,
        'classes': classes,
        'students': students,
        'search_teacher': search_teacher,
        'search_class': search_class,
        'search_student': search_student,
    }
    
    return render_template('admin_dashboard.html', **context)


@app.route('/teacher/dashboard')
@teacher_required
def teacher_dashboard():
    """Teacher dashboard - checks teacher_id session variable ONLY"""
    # Get teacher user from session
    teacher_user = db.session.get(User, session.get('teacher_id'))
    teacher = Teacher.query.filter_by(user_id=teacher_user.id).first()
    
    if not teacher:
        flash('Teacher profile not found.', 'danger')
        return redirect(url_for('teacher_logout'))
    
    # Get search query
    search_query = request.args.get('search', '').strip()
    
    # Show ALL classes to all teachers
    classes = Class.query.all()
    
    # Filter classes if search query provided (by course, specialisation, or teacher name)
    if search_query:
        classes = [cls for cls in classes if 
                  search_query.lower() in cls.course.lower() or 
                  search_query.lower() in cls.specialisation.lower() or
                  search_query.lower() in (cls.teacher.full_name.lower() if cls.teacher else '')]
    
    context = {
        'teacher': teacher,
        'classes': classes,
        'classes_count': len(classes),
        'students_count': sum([len(cls.students) for cls in classes]),
        'search_query': search_query,
    }
    
    return render_template('teacher_dashboard.html', **context)


# ==================== Admin - Teacher Management ====================

@app.route('/admin/teachers/add', methods=['GET', 'POST'])
@admin_required
def add_teacher():
    """Admin: Add new teacher account"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        full_name = request.form.get('full_name', '').strip()
        email = request.form.get('email', '').strip()
        phone = request.form.get('phone', '').strip()
        department = request.form.get('department', '').strip()
        
        # Validation
        if not username or not password:
            flash('Username and password are required.', 'danger')
            return redirect(url_for('add_teacher'))
        
        if User.query.filter_by(username=username).first():
            flash('Username already exists.', 'danger')
            return redirect(url_for('add_teacher'))
        
        try:
            # Create user
            user = User(username=username, role='teacher')
            user.set_password(password)
            db.session.add(user)
            db.session.flush()
            
            # Create teacher profile
            teacher = Teacher(
                user_id=user.id,
                full_name=full_name,
                email=email,
                phone=phone,
                department=department
            )
            db.session.add(teacher)
            db.session.commit()
            
            flash(f'Teacher account created successfully: {username}', 'success')
            return redirect(url_for('admin_dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating teacher: {str(e)}', 'danger')
    
    return render_template('add_teacher.html')


@app.route('/admin/teachers')
@admin_required
def view_teachers():
    """Admin: View all teachers"""
    teachers = Teacher.query.all()
    return render_template('view_teachers.html', teachers=teachers)


@app.route('/admin/teacher/<int:teacher_id>')
@admin_required
def view_teacher(teacher_id):
    """Admin: View teacher details"""
    teacher = Teacher.query.get_or_404(teacher_id)
    classes = teacher.classes
    total_students = sum(len(list(cls.students)) for cls in classes) if classes else 0
    
    context = {
        'teacher': teacher,
        'classes': classes,
        'total_students': total_students,
    }
    
    return render_template('view_teacher.html', **context)


@app.route('/admin/teacher/<int:teacher_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_teacher(teacher_id):
    """Admin: Edit teacher details"""
    teacher = Teacher.query.get_or_404(teacher_id)
    
    if request.method == 'POST':
        teacher.full_name = request.form.get('full_name', '').strip()
        teacher.email = request.form.get('email', '').strip()
        teacher.phone = request.form.get('phone', '').strip()
        teacher.department = request.form.get('department', '').strip()
        
        try:
            db.session.commit()
            flash('Teacher updated successfully.', 'success')
            return redirect(url_for('view_teacher', teacher_id=teacher.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating teacher: {str(e)}', 'danger')
    
    return render_template('edit_teacher.html', teacher=teacher)


@app.route('/admin/teacher/<int:teacher_id>/delete', methods=['POST'])
@admin_required
def delete_teacher(teacher_id):
    """Admin: Delete teacher"""
    teacher = Teacher.query.get_or_404(teacher_id)
    user_id = teacher.user_id
    
    try:
        db.session.delete(teacher)
        user = db.session.get(User, user_id)
        if user:
            db.session.delete(user)
        db.session.commit()
        flash('Teacher deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting teacher: {str(e)}', 'danger')
    
    return redirect(url_for('view_teachers'))


# ==================== Class Management ====================

@app.route('/teacher/class/add', methods=['GET', 'POST'])
@teacher_required
def add_class():
    """Teacher: Add new class"""
    if request.method == 'POST':
        course = request.form.get('course', '').strip()
        specialisation = request.form.get('specialisation', '').strip()
        
        if not course or not specialisation:
            flash('Course and specialisation are required.', 'danger')
            return redirect(url_for('add_class'))
        
        # Check if class already exists (same course + specialisation)
        existing_class = Class.query.filter_by(
            course=course,
            specialisation=specialisation
        ).first()
        
        if existing_class:
            flash(f'Class "{course} - {specialisation}" already exists! Created by {existing_class.teacher.full_name if existing_class.teacher else "Unknown"}', 'warning')
            return redirect(url_for('teacher_dashboard'))
        
        try:
            # Get teacher from session (teacher_id is set when teacher logs in)
            teacher = Teacher.query.filter_by(user_id=session['teacher_id']).first()
            
            if not teacher:
                flash('Teacher profile not found.', 'danger')
                return redirect(url_for('teacher_dashboard'))
            
            class_obj = Class(
                teacher_id=teacher.id,
                course=course,
                specialisation=specialisation
            )
            db.session.add(class_obj)
            db.session.commit()
            
            flash('Class added successfully.', 'success')
            return redirect(url_for('teacher_dashboard'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding class: {str(e)}', 'danger')
    
    return render_template('add_class.html')


@app.route('/class/<int:class_id>/dashboard')
@login_required
def class_dashboard(class_id):
    """Class dashboard - view and manage students"""
    class_obj = Class.query.get_or_404(class_id)
    students = list(class_obj.students)
    
    # Get search query
    search_query = request.args.get('search', '').strip()
    
    # Filter students if search query provided (by USN or name)
    if search_query:
        students = [s for s in students if 
                   search_query.lower() in s.usn.lower() or
                   search_query.lower() in s.full_name.lower()]
        
        # Remove duplicates while preserving order
        seen_usn = set()
        unique_students = []
        for student in students:
            if student.usn not in seen_usn:
                unique_students.append(student)
                seen_usn.add(student.usn)
        students = unique_students
    else:
        # Remove duplicates from all students too (safety check)
        seen_usn = set()
        unique_students = []
        for student in students:
            if student.usn not in seen_usn:
                unique_students.append(student)
                seen_usn.add(student.usn)
        students = unique_students
    
    context = {
        'class': class_obj,
        'students': students,
        'students_count': len(students),
        'placed_count': len([s for s in students if s.company_name]),
        'search_query': search_query,
    }
    
    return render_template('class_dashboard.html', **context)


# --- Excel import helpers and teacher endpoints (Bulk upload, preview, update) --- #

import csv
import time
import uuid
from pathlib import Path

EXCEL_UPLOAD_DIR = os.path.join(UPLOAD_FOLDER, 'excel_uploads')
Path(EXCEL_UPLOAD_DIR).mkdir(parents=True, exist_ok=True)

# Delay (seconds) between notifications to avoid rate limits and ensure sequential calls
# Delay (seconds) between notifications to avoid rate limits and ensure sequential calls
EXCEL_NOTIFICATION_DELAY_SECONDS = int(os.getenv('EXCEL_NOTIFICATION_DELAY_SECONDS', '3'))
# Number of retries for call attempts
EXCEL_CALL_RETRY_COUNT = int(os.getenv('EXCEL_CALL_RETRY_COUNT', '2'))

# Background task rate-limits
CALL_RATE_DELAY_SECONDS = int(os.getenv('CALL_RATE_DELAY_SECONDS', '10'))  # seconds between calls
WHATSAPP_RATE_DELAY_SECONDS = int(os.getenv('WHATSAPP_RATE_DELAY_SECONDS', '1'))  # seconds between whatsapp sends in worker

# Background queues
CALL_TASK_QUEUE = Queue()
WHATSAPP_TASK_QUEUE = Queue()
CLIENT_QUEUES = set()
# Number of retries for call attempts
EXCEL_CALL_RETRY_COUNT = int(os.getenv('EXCEL_CALL_RETRY_COUNT', '2'))


def _save_uploaded_excel(file_obj, prefix='excel'):
    """Save uploaded file to instance upload folder and return filepath"""
    filename = secure_filename(f"{prefix}_{int(time.time())}_{file_obj.filename}")
    filepath = os.path.join(EXCEL_UPLOAD_DIR, filename)
    file_obj.save(filepath)
    return filepath


def _parse_excel_file(filepath):
    """
    Parse a CSV/XLSX/XLS into list[dict].
    Supports: CSV (utf-8) and XLSX via openpyxl (no mandatory pandas dependency).
    """
    ext = filepath.rsplit('.', 1)[-1].lower()
    rows = []
    try:
        if ext in ('xlsx', 'xls'):
            try:
                import openpyxl
            except ImportError:
                raise RuntimeError("openpyxl required to read Excel files (pip install openpyxl)")
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            it = ws.iter_rows(values_only=True)
            headers = []
            for i, row in enumerate(it):
                if i == 0:
                    headers = [str(cell).strip() if cell is not None else '' for cell in row]
                    continue
                if not any(cell is not None and str(cell).strip() != '' for cell in row):
                    continue
                d = {}
                for h, cell in zip(headers, row):
                    if h:
                        d[h.strip().lower()] = str(cell).strip() if cell is not None else ''
                rows.append(d)
        elif ext == 'csv':
            with open(filepath, newline='', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    normalized = {}
                    for k, v in row.items():
                        if k is None:
                            # Skip any extra values without corresponding headers
                            continue
                        kk = k.strip().lower()
                        normalized[kk] = (v.strip() if v else '')
                    rows.append(normalized)
        else:
            raise RuntimeError('Unsupported file type')
    except Exception as e:
        return {'error': str(e)}
    return rows


def _normalize_row_keys(row):
    """Helper to map ambiguous Excel column names to our expected fields"""
    mapping = {
        'usn': 'usn',
        'usn number': 'usn',
        'name': 'full_name',
        'full_name': 'full_name',
        'email': 'email_id',
        'email_id': 'email_id',
        'phone': 'phone_number',
        'phone number': 'phone_number',
        'phone_number': 'phone_number',
        'mother tongue': 'mother_tongue',
        'mother_tongue': 'mother_tongue',
        'parent name': 'parent_name',
        'parent_name': 'parent_name',
        'parent email': 'parent_email',
        'parent_email': 'parent_email',
        'parent phone': 'parent_phone',
        'parent_phone': 'parent_phone',
        'company': 'company_name',
        'company name': 'company_name',
        'company_name': 'company_name',
        'package': 'package',
        'package (lpa)': 'package',
       
    }
    out = {}
    for k, v in row.items():
        if not k:
            continue
        kk = k.strip().lower()
        mapped = mapping.get(kk, kk)
        out[mapped] = v.strip() if isinstance(v, str) else (str(v).strip() if v is not None else '')
    return out


def _compute_preview(class_id, rows):
    """Compare Excel rows with DB, return preview (to_insert, to_update, to_deactivate)"""
    existing = {s.usn.upper(): s for s in Student.query.filter_by(class_id=class_id).all()}
    usns_in_file = set()
    to_insert = []
    to_update = []
    for r in rows:
        rnorm = _normalize_row_keys(r)
        usn = (rnorm.get('usn') or '').strip().upper()
        if not usn:
            continue
        usns_in_file.add(usn)
        st = existing.get(usn)
        if st:
            # detect changes
            changes = {}
            for fld in ['full_name', 'phone_number', 'email_id', 'mother_tongue',
                        'parent_name', 'parent_phone', 'parent_email', 'company_name',
                        'package', 'preferred_voice', 'preferred_prosody', 'external_id']:
                newv = rnorm.get(fld) or None
                oldv = getattr(st, fld)
                if (str(newv or '') != str(oldv or '')):
                    changes[fld] = {'old': oldv, 'new': newv}
            if changes:
                to_update.append({'usn': usn, 'changes': changes, 'data': rnorm})
        else:
            to_insert.append({'usn': usn, 'data': rnorm})
    # Deactivate: any existing usn not in the file
    to_deactivate = [usn for usn in existing.keys() if usn not in usns_in_file]
    return {
        'to_insert_count': len(to_insert),
        'to_update_count': len(to_update),
        'to_deactivate_count': len(to_deactivate),
        'to_insert': to_insert[:50],
        'to_update': to_update[:50],
        'to_deactivate': to_deactivate[:50],
    }


def _apply_updates(class_id, rows, notify=True):
    """
    Apply inserts / updates / deactivations to DB.
    Returns list of per-student results including notification statuses when notify=True.
    """
    results = []
    rows_norm = [_normalize_row_keys(r) for r in rows]
    existing = {s.usn.upper(): s for s in Student.query.filter_by(class_id=class_id).all()}
    seen_usns = set()

    # Insert / Update from file
    for r in rows_norm:
        usn = (r.get('usn') or '').strip().upper()
        if not usn:
            continue
        seen_usns.add(usn)
        student = existing.get(usn)
        if student is None:
            # Insert
            student = Student(
                class_id=class_id,
                usn=usn,
                full_name=r.get('full_name') or '',
                phone_number=r.get('phone_number') or '',
                email_id=r.get('email_id') or '',
                mother_tongue=r.get('mother_tongue') or '',
                parent_name=r.get('parent_name') or '',
                parent_phone=r.get('parent_phone') or '',
                parent_email=r.get('parent_email') or '',
                company_name=r.get('company_name') or None,
                package=r.get('package') or None,
               
            )
            db.session.add(student)
            action = 'inserted'
        else:
            # Update
            action = 'updated'
            for fld in ['full_name', 'phone_number', 'email_id', 'mother_tongue',
                        'parent_name', 'parent_phone', 'parent_email', 'company_name',
                        'package']:
                val = r.get(fld)
                if val is not None:
                    setattr(student, fld, val or None)
            student.is_active = (str(r.get('is_active') or 'true').strip().lower() not in ['false', '0', 'no'])

        try:
            db.session.flush()  # get id for new rows
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            results.append({'usn': usn, 'action': action, 'success': False, 'error': str(e)})
            continue

        # If company+package present -> trigger notification (if requested)
        notify_result = None
        if notify and student.company_name and student.package:
            try:
                # Send WhatsApp + voice + call where applicable and update fields
                from datetime import datetime
                student_data = student.to_dict()

                print(f"[BULK] Sending notifications for USN {student.usn} - {student.full_name}")

                # WhatsApp (blocking call)
                whatsapp_result = send_placement_whatsapp(student_data)

                # Call: enqueue for background processing to avoid overloading Twilio
                pn = student.parent_phone
                call_script = get_placement_notifications(student.full_name, student.company_name, student.package, student.mother_tongue).get('call_script', '')
                enqueue_call_task(student.id)
                call_result = {'success': None, 'message': 'queued'}
                # mark queued state in DB; worker will update to sent/failed
                student.last_call_status = 'queued'

                # Update DB statuses
                student.whatsapp_notification_sent = bool(whatsapp_result.get('success') if isinstance(whatsapp_result, dict) else whatsapp_result)
                student.last_whatsapp_status = 'sent' if student.whatsapp_notification_sent else 'failed'
                student.last_call_status = 'sent' if call_result.get('success') else 'failed'
                student.notification_sent = student.whatsapp_notification_sent
                student.notification_sent_at = datetime.utcnow() if student.whatsapp_notification_sent else student.notification_sent_at
                db.session.commit()
                notify_result = {'whatsapp': whatsapp_result, 'call': call_result}
                # Broadcast student updated to clients
                try:
                    broadcast_event({'type': 'student_updated', 'student_id': student.id, 'class_id': class_id})
                except Exception:
                    pass
            except Exception as e:
                db.session.rollback()
                notify_result = {'error': str(e)}
        results.append({'usn': usn, 'action': action, 'success': True, 'notify': notify_result})
        # Small pause to avoid rate limits and ensure sequential behaviour
        time.sleep(EXCEL_NOTIFICATION_DELAY_SECONDS)

    # Deactivate any students missing from file
    all_usns_in_db = set(existing.keys())
    to_deactivate = [usn for usn in all_usns_in_db if usn not in seen_usns]
    for usn in to_deactivate:
        st = existing.get(usn)
        if st and st.is_active:
            st.is_active = False
            try:
                db.session.commit()
                try:
                    broadcast_event({'type': 'student_updated', 'student_id': st.id, 'class_id': class_id})
                except Exception:
                    pass
                results.append({'usn': usn, 'action': 'deactivated', 'success': True})
            except Exception as e:
                db.session.rollback()
                results.append({'usn': usn, 'action': 'deactivated', 'success': False, 'error': str(e)})

    return results


def enqueue_call_task(student_id):
    try:
        CALL_TASK_QUEUE.put(student_id)
        print(f"[QUEUE] Enqueued call for student {student_id}")
        # Mark queued state in DB so UI shows queued
        try:
            with app.app_context():
                student = Student.query.get(student_id)
                if student and student.last_call_status != 'queued':
                    student.last_call_status = 'queued'
                    db.session.commit()
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass
    except Exception as e:
        print(f"[QUEUE] Failed to enqueue call: {e}")
        return False
        # Broadcast queue event to clients
        broadcast_event({'type': 'task_queued', 'task': 'call', 'student_id': student_id})
        return True


def enqueue_whatsapp_task(student_id):
    try:
        WHATSAPP_TASK_QUEUE.put(student_id)
        print(f"[QUEUE] Enqueued whatsapp for student {student_id}")
        # Mark queued state in DB so UI shows queued
        try:
            with app.app_context():
                student = Student.query.get(student_id)
                if student and student.last_whatsapp_status != 'queued':
                    student.last_whatsapp_status = 'queued'
                    db.session.commit()
        except Exception:
            try:
                db.session.rollback()
            except Exception:
                pass
    except Exception as e:
        print(f"[QUEUE] Failed to enqueue whatsapp: {e}")
        return False
        # Broadcast queue event to clients
        broadcast_event({'type': 'task_queued', 'task': 'whatsapp', 'student_id': student_id})
        return True


def call_worker():
    print("[WORKER] Call worker thread started")
    while True:
        student_id = CALL_TASK_QUEUE.get()
        try:
            with app.app_context():
                student = Student.query.get(student_id)
                if not student:
                    print(f"[WORKER] Student {student_id} not found for call task")
                    continue
                print(f"[WORKER] Processing call for {student.usn} -> {student.parent_phone}")
                student_data = student.to_dict()
                placement_notifications = get_placement_notifications(student.full_name, student.company_name, student.package, student.mother_tongue)
                call_script = placement_notifications.get('call_script', '')
                language_code = placement_notifications.get('language_code', 'en')
                call_result = send_placement_call_with_twilio(
                    student.parent_phone,
                    call_script,
                    student.full_name,
                    student.company_name,
                    language_code,
                    call_voice=student.preferred_voice,
                    call_prosody=student.preferred_prosody
                )
                student.last_call_status = 'sent' if (call_result.get('success') if isinstance(call_result, dict) else call_result) else 'failed'
                db.session.commit()
                print(f"[WORKER] Call result for {student.usn}: {student.last_call_status}")
                broadcast_event({'type': 'task_completed', 'task': 'call', 'student_id': student.id, 'status': student.last_call_status, 'last_call_status': student.last_call_status})
        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            print(f"[WORKER] Error processing call task for {student_id}: {e}")
        finally:
            try:
                time.sleep(CALL_RATE_DELAY_SECONDS)
            except Exception:
                pass
            CALL_TASK_QUEUE.task_done()


def broadcast_event(payload: dict):
    """Send payload dict (JSON) to all connected SSE clients by pushing to their queues."""
    import json
    p = json.dumps(payload)
    dead_clients = []
    for q in list(CLIENT_QUEUES):
        try:
            q.put(p)
        except Exception as e:
            print(f"[SSE] Failed to broadcast to client queue: {e}")
            dead_clients.append(q)
    # Remove dead clients
    for q in dead_clients:
        CLIENT_QUEUES.discard(q)


def whatsapp_worker():
    print("[WORKER] WhatsApp worker thread started")
    while True:
        student_id = WHATSAPP_TASK_QUEUE.get()
        try:
            with app.app_context():
                student = Student.query.get(student_id)
                if not student:
                    print(f"[WORKER] Student {student_id} not found for whatsapp task")
                    continue
                print(f"[WORKER] Sending whatsapp to {student.usn} -> {student.parent_phone}")
                whatsapp_result = send_placement_whatsapp(student.to_dict())
                whatsapp_success = (whatsapp_result.get('success') if isinstance(whatsapp_result, dict) else whatsapp_result)
                student.whatsapp_notification_sent = bool(whatsapp_success)
                student.last_whatsapp_status = 'sent' if whatsapp_success else 'failed'
                if whatsapp_success:
                    student.notification_sent = True
                    student.notification_sent_at = datetime.utcnow()
                db.session.commit()
                print(f"[WORKER] WhatsApp result for {student.usn}: {student.last_whatsapp_status}")
                broadcast_event({'type': 'task_completed', 'task': 'whatsapp', 'student_id': student.id, 'status': student.last_whatsapp_status, 'whatsapp_notification_sent': student.whatsapp_notification_sent})
        except Exception as e:
            try:
                db.session.rollback()
            except Exception:
                pass
            print(f"[WORKER] Error processing whatsapp task for {student_id}: {e}")
        finally:
            try:
                time.sleep(WHATSAPP_RATE_DELAY_SECONDS)
            except Exception:
                pass
            WHATSAPP_TASK_QUEUE.task_done()


def start_task_workers():
    # start workers if they are not already started
    if not hasattr(app, 'task_workers_started') or not app.task_workers_started:
        t1 = threading.Thread(target=call_worker, daemon=True)
        t1.start()
        t2 = threading.Thread(target=whatsapp_worker, daemon=True)
        t2.start()
        app.task_workers_started = True
        print('[WORKER] Background task workers started')


@app.route('/events')
def sse_events():
    """Server-Sent Events endpoint used by clients to receive background updates."""
    from flask import Response, stream_with_context
    import json

    def gen():
        q = Queue()
        CLIENT_QUEUES.add(q)
        try:
            while True:
                try:
                    data = q.get(timeout=15)
                except Exception:
                    # heartbeat
                    yield "event: ping\ndata: {}\n\n"
                    continue
                yield f"data: {data}\n\n"
        finally:
            CLIENT_QUEUES.discard(q)

    return Response(stream_with_context(gen()), mimetype='text/event-stream')


@app.route('/class/<int:class_id>/student-row/<int:student_id>')
def get_student_row(class_id, student_id):
    from flask import abort
    student = Student.query.filter_by(id=student_id, class_id=class_id).first()
    if not student:
        abort(404)
    return render_template('_student_row.html', student=student)


@app.route('/teacher/<int:class_id>/upload_excel', methods=['POST'])
@teacher_required
def teacher_upload_excel(class_id):
    """Upload Excel file for preview/update (saves file and returns saved filename)."""
    if 'excel_file' not in request.files:
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400
    f = request.files['excel_file']
    if not f or f.filename == '':
        return jsonify({'success': False, 'message': 'No file uploaded'}), 400
    ext = f.filename.rsplit('.', 1)[-1].lower()
    if ext not in ['xlsx', 'xls', 'csv']:
        return jsonify({'success': False, 'message': 'Unsupported file type'}), 400
    filepath = _save_uploaded_excel(f, prefix=f"class{class_id}")
    return jsonify({'success': True, 'filename': os.path.basename(filepath), 'path': filepath})


@app.route('/teacher/<int:class_id>/preview_excel', methods=['POST'])
@teacher_required
def teacher_preview_excel(class_id):
    """
    Parse uploaded Excel (file or saved filename) and return preview:
    to_insert_count, to_update_count, to_deactivate_count and sample rows.
    """
    # Accept either a direct file upload or a previously saved filename
    file_obj = request.files.get('excel_file')
    if file_obj:
        filepath = _save_uploaded_excel(file_obj, prefix=f"class{class_id}_preview")
    else:
        filename = request.form.get('filename') or request.json.get('filename') if request.json else None
        if not filename:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        filepath = os.path.join(EXCEL_UPLOAD_DIR, filename)
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': 'File not found on server'}), 404

    parsed = _parse_excel_file(filepath)
    if isinstance(parsed, dict) and parsed.get('error'):
        return jsonify({'success': False, 'message': parsed.get('error')}), 400
    # Basic header validation: ensure required columns are present (USN, Name, Phone number)
    if parsed and isinstance(parsed, list) and len(parsed) > 0:
        # Build a fake row from headers to test normalization mapping
        header_keys = [k for k in parsed[0].keys() if k]
        fake_row = {k: '' for k in header_keys}
        normalized_headers = set(_normalize_row_keys(fake_row).keys())
        required_map = {'usn': 'USN', 'full_name': 'Name', 'phone_number': 'Phone number'}
        missing = [v for k, v in required_map.items() if k not in normalized_headers]
        if missing:
            return jsonify({'success': False, 'message': f'Missing required columns: {", ".join(missing)}'}), 400
    preview = _compute_preview(class_id, parsed)
    # Provide a lightweight sample for UI
    sample_rows = parsed[:50]
    return jsonify({'success': True, 'preview': preview, 'sample': sample_rows, 'filename': os.path.basename(filepath)})


@app.route('/teacher/<int:class_id>/download_excel_template')
@teacher_required
def download_excel_template(class_id):
    """Generate and return a blank Excel/CSV template for this class (headers only)."""
    # Define canonical headers used by the import
    headers = [
        'USN', 'Name', 'Email', 'Phone number', 'Mother Tongue', 'Parent Name', 'Parent Email', 'Parent Phone', 'Company Name', 'Package (LPA)'
    ]

    fmt = request.args.get('format', 'xlsx').lower()
    filename = f'class_{class_id}_template.{"csv" if fmt == "csv" else "xlsx"}'

    if fmt == 'csv':
        # CSV fallback
        import csv
        sio = io.StringIO()
        writer = csv.DictWriter(sio, fieldnames=headers)
        writer.writeheader()
        writer.writerow({h: '' for h in headers})
        data = sio.getvalue().encode('utf-8')
        return send_file(io.BytesIO(data), as_attachment=True, download_name=filename, mimetype='text/csv')

    # Default: XLSX using openpyxl if available
    try:
        import openpyxl
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(['' for _ in headers])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except ImportError:
        # Fallback to CSV if openpyxl isn't installed
        import csv
        sio = io.StringIO()
        writer = csv.DictWriter(sio, fieldnames=headers)
        writer.writeheader()
        writer.writerow({h: '' for h in headers})
        data = sio.getvalue().encode('utf-8')
        return send_file(io.BytesIO(data), as_attachment=True, download_name=f'class_{class_id}_template.csv', mimetype='text/csv')


@app.route('/teacher/<int:class_id>/update_excel', methods=['POST'])
@teacher_required
def teacher_update_excel(class_id):
    """
    Apply an uploaded Excel file to insert/update/deactivate students in this class.
    JSON/form params:
        - excel_file (file) or filename (string) referring to previously uploaded file
        - notify (boolean: default True) -> whether to send notifications for rows with company+package
    """
    # Retrieve file
    file_obj = request.files.get('excel_file')
    if file_obj:
        filepath = _save_uploaded_excel(file_obj, prefix=f"class{class_id}_update")
    else:
        data = request.get_json() or request.form
        filename = data.get('filename')
        if not filename:
            return jsonify({'success': False, 'message': 'No file provided'}), 400
        filepath = os.path.join(EXCEL_UPLOAD_DIR, filename)
        if not os.path.exists(filepath):
            return jsonify({'success': False, 'message': 'File not found'}), 404

    notify = request.form.get('notify', '1') != '0' if request.form else request.get_json().get('notify', True)

    parsed = _parse_excel_file(filepath)
    if isinstance(parsed, dict) and parsed.get('error'):
        return jsonify({'success': False, 'message': parsed.get('error')}), 400

    # Apply updates (this commits per-row)
    results = _apply_updates(class_id, parsed, notify=bool(notify))
    summary = {
        'total_rows': len(parsed),
        'actions': {
            'inserted': sum(1 for r in results if r.get('action') == 'inserted' and r.get('success')),
            'updated': sum(1 for r in results if r.get('action') == 'updated' and r.get('success')),
            'deactivated': sum(1 for r in results if r.get('action') == 'deactivated' and r.get('success')),
            'failed': sum(1 for r in results if not r.get('success')),
        }
    }
    return jsonify({'success': True, 'summary': summary, 'details': results})


@app.route('/class/<int:class_id>/edit', methods=['GET', 'POST'])
@teacher_required
def edit_class(class_id):
    """Teacher/Admin: Edit class"""
    class_obj = Class.query.get_or_404(class_id)
    
    if request.method == 'POST':
        class_obj.course = request.form.get('course', '').strip()
        class_obj.specialisation = request.form.get('specialisation', '').strip()
        
        try:
            db.session.commit()
            flash('Class updated successfully.', 'success')
            return redirect(url_for('class_dashboard', class_id=class_obj.id))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating class: {str(e)}', 'danger')
    
    return render_template('edit_class.html', class_obj=class_obj)


@app.route('/class/<int:class_id>/delete', methods=['POST'])
@teacher_required
def delete_class(class_id):
    """Teacher: Delete class"""
    class_obj = Class.query.get_or_404(class_id)
    
    try:
        db.session.delete(class_obj)
        db.session.commit()
        flash('Class deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting class: {str(e)}', 'danger')
    
    return redirect(url_for('teacher_dashboard'))


# ==================== Student Management ====================

@app.route('/class/<int:class_id>/student/add', methods=['GET', 'POST'])
@login_required
def add_student(class_id):
    """Add new student to class"""
    class_obj = Class.query.get_or_404(class_id)
    
    if request.method == 'POST':
        # Get form data
        usn = request.form.get('usn', '').strip().upper()
        full_name = request.form.get('full_name', '').strip()
        phone_number = request.form.get('phone_number', '').strip()
        email_id = request.form.get('email_id', '').strip()
        mother_tongue = request.form.get('mother_tongue', '').strip()
        parent_name = request.form.get('parent_name', '').strip()
        parent_phone = request.form.get('parent_phone', '').strip()
        parent_email = request.form.get('parent_email', '').strip()
        company_name = request.form.get('company_name', '').strip() or None
        package = request.form.get('package', '').strip() or None
        send_whatsapp = request.form.get('send_whatsapp') == '1'
        
        # Validation
        if not all([usn, full_name, phone_number, email_id, parent_name, parent_phone, parent_email]):
            flash('All required fields must be filled.', 'danger')
            return redirect(url_for('add_student', class_id=class_id))
        
        # Check for duplicate USN
        if Student.query.filter_by(usn=usn).first():
            flash('USN already exists.', 'danger')
            return redirect(url_for('add_student', class_id=class_id))
        
        try:
            student = Student(
                class_id=class_id,
                usn=usn,
                full_name=full_name,
                phone_number=phone_number,
                email_id=email_id,
                mother_tongue=mother_tongue,
                parent_name=parent_name,
                parent_phone=parent_phone,
                parent_email=parent_email,
                company_name=company_name,
                package=package,
                send_whatsapp=send_whatsapp
            )
            
            # Handle file upload
            if 'confirmation_letter' in request.files:
                file = request.files['confirmation_letter']
                if file and file.filename and allowed_file(file.filename):
                    filename = secure_filename(f"{usn}_{file.filename}")
                    filepath = os.path.join(UPLOAD_FOLDER, filename)
                    file.save(filepath)
                    student.confirmation_letter_path = filepath
            
            db.session.add(student)
            db.session.commit()
            
            # Trigger placement notification if both company and package are filled (queued)
            if company_name and package:
                notification_sent = trigger_placement_notification(student)
                if notification_sent:
                    flash('Student added and placement notification queued!', 'success')
                else:
                    flash('Student added. Placement notification could not be queued.', 'info')
            # Broadcast student added
            try:
                broadcast_event({'type': 'student_updated', 'student_id': student.id, 'class_id': class_id})
            except Exception:
                pass
            else:
                flash('Student added successfully.', 'success')
            
            return redirect(url_for('class_dashboard', class_id=class_id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error adding student: {str(e)}', 'danger')
    
    return render_template('add_student.html', class_obj=class_obj)


@app.route('/student/<int:student_id>')
@login_required
def view_student(student_id):
    """View student details"""
    student = Student.query.get_or_404(student_id)
    return render_template('view_student.html', student=student)


@app.route('/student/<int:student_id>/edit', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    """Edit student details"""
    student = Student.query.get_or_404(student_id)
    
    if request.method == 'POST':
        # Update student fields
        student.full_name = request.form.get('full_name', '').strip()
        student.phone_number = request.form.get('phone_number', '').strip()
        student.email_id = request.form.get('email_id', '').strip()
        student.mother_tongue = request.form.get('mother_tongue', '').strip()
        student.parent_name = request.form.get('parent_name', '').strip()
        student.parent_phone = request.form.get('parent_phone', '').strip()
        student.parent_email = request.form.get('parent_email', '').strip()
        student.send_whatsapp = request.form.get('send_whatsapp') == '1'
        
        old_company = student.company_name
        old_package = student.package
        
        student.company_name = request.form.get('company_name', '').strip() or None
        student.package = request.form.get('package', '').strip() or None
        
        # Handle file upload
        if 'confirmation_letter' in request.files:
            file = request.files['confirmation_letter']
            if file and file.filename and allowed_file(file.filename):
                filename = secure_filename(f"{student.usn}_{file.filename}")
                filepath = os.path.join(UPLOAD_FOLDER, filename)
                file.save(filepath)
                student.confirmation_letter_path = filepath
        
        try:
            db.session.commit()
            
            # Trigger notification if company/package just filled - schedule in background
            if (not old_company or not old_package) and student.company_name and student.package:
                # mark queued states so UI reflects the task
                student.last_whatsapp_status = 'queued'
                student.last_call_status = 'queued'
                try:
                    db.session.commit()
                except Exception:
                    db.session.rollback()
                enqueue_whatsapp_task(student.id)
                enqueue_call_task(student.id)
                flash('Student updated and placement notification queued!', 'success')
            # Broadcast update to clients
            try:
                broadcast_event({'type': 'student_updated', 'student_id': student.id, 'class_id': student.class_id})
            except Exception:
                pass
            else:
                flash('Student updated successfully.', 'success')
            
            return redirect(url_for('view_student', student_id=student.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating student: {str(e)}', 'danger')
    
    return render_template('edit_student.html', student=student)


@app.route('/student/<int:student_id>/delete', methods=['POST'])
@login_required
def delete_student(student_id):
    """Delete student"""
    student = Student.query.get_or_404(student_id)
    class_id = student.class_id
    
    try:
        # Delete uploaded file if exists
        if student.confirmation_letter_path and os.path.exists(student.confirmation_letter_path):
            os.remove(student.confirmation_letter_path)
        
        db.session.delete(student)
        db.session.commit()
        flash('Student deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting student: {str(e)}', 'danger')
    
    return redirect(url_for('class_dashboard', class_id=class_id))


# ==================== WhatsApp Notifications ====================

@app.route('/initialize-whatsapp', methods=['POST'])
@login_required
def initialize_whatsapp_route():
    """Initialize WhatsApp Web connection"""
    try:
        print("\n" + "="*60)
        print("[*] Initializing WhatsApp Web...")
        print("="*60)
        
        # This will open WhatsApp Web and guide user through QR code
        initialize_whatsapp()
        
        return jsonify({
            'success': True,
            'message': 'WhatsApp Web initialized. Check your browser for QR code.'
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': f'Error initializing WhatsApp: {str(e)}'
        }), 500


@app.route('/send-whatsapp-notification', methods=['POST'])
@login_required
def send_whatsapp_notification():
    """Send WhatsApp notification to a student"""
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        
        student = Student.query.get_or_404(student_id)
        
        # Check if student has both company and package
        if not student.company_name or not student.package:
            return jsonify({
                'success': False,
                'message': 'Student must have company and package details filled'
            }), 400
        
        # Prepare student data
        student_data = student.to_dict()
        
        # Generate voice message (with error handling for Windows permission issues)
        
        # Note: Voice generation now happens automatically inside send_placement_whatsapp()
        # with translation support, so we don't pre-generate it here anymore
        
        # Send WhatsApp notification
        whatsapp_result = send_placement_whatsapp(student_data)

        sent_success = whatsapp_result.get('success') if isinstance(whatsapp_result, dict) else whatsapp_result
        db_updated = False
        if sent_success:
            try:
                student.whatsapp_notification_sent = True
                student.notification_sent_at = datetime.utcnow()
                db.session.commit()
                db_updated = True
            except Exception as commit_err:
                db.session.rollback()
                print(f"Failed to update DB after sending WhatsApp: {commit_err}")

        # Broadcast task_completed so dashboard updates immediately
        try:
            broadcast_event({'type': 'task_completed', 'task': 'whatsapp', 'student_id': student.id, 'status': ('sent' if sent_success else 'failed'), 'whatsapp_notification_sent': db_updated})
        except Exception:
            pass

        # Provide detailed result to the client so it can decide on recovery
        return jsonify({
            'success': bool(sent_success),
            'message': 'WhatsApp message sent' if sent_success else 'Failed to send WhatsApp message',
            'whatsapp_result': whatsapp_result if isinstance(whatsapp_result, dict) else {'success': bool(whatsapp_result)},
            'db_updated': db_updated
        }), (200 if sent_success else 500)
    
    except Exception as e:
        db.session.rollback()
        print(f"Error sending WhatsApp notification: {e}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500


@app.route('/mark-whatsapp-sent', methods=['POST'])
@login_required
def mark_whatsapp_sent():
    """Mark a student's whatsapp_notification_sent flag in DB (idempotent)"""
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        student = Student.query.get_or_404(student_id)
        if not student.whatsapp_notification_sent:
            student.whatsapp_notification_sent = True
            student.notification_sent_at = datetime.utcnow()
            db.session.commit()
            try:
                broadcast_event({'type': 'task_completed', 'task': 'whatsapp', 'student_id': student.id, 'status': 'sent', 'whatsapp_notification_sent': True})
            except Exception:
                pass
        return jsonify({'success': True, 'message': 'Marked as sent'})
    except Exception as e:
        db.session.rollback()
        print(f"Error marking WhatsApp sent in DB: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ==================== Admin View Routes ====================

@app.route('/admin/classes')
@admin_required
def admin_view_classes():
    """Admin: View all classes"""
    classes = Class.query.all()
    return render_template('admin_view_classes.html', classes=classes)


@app.route('/admin/class/<int:class_id>')
@admin_required
def admin_view_class(class_id):
    """Admin: View class details"""
    class_obj = Class.query.get_or_404(class_id)
    return redirect(url_for('class_dashboard', class_id=class_id))


@app.route('/admin/students')
@admin_required
def admin_view_students():
    """Admin: View all students"""
    students = Student.query.all()
    return render_template('admin_view_students.html', students=students)


# ==================== Error Handlers ====================

@app.errorhandler(404)
def page_not_found(error):
    """Handle 404 errors"""
    return render_template('404.html'), 404


@app.errorhandler(500)
def internal_error(error):
    """Handle 500 errors"""
    db.session.rollback()
    return render_template('500.html'), 500


# ==================== Main ====================

if __name__ == '__main__':
    with app.app_context():
        # Check if database exists, create if not
        db_path = os.path.join(app.instance_path, 'database.db')
        db_exists = os.path.exists(db_path)
        
        # Create all tables
        db.create_all()
        
        # Create default admin user
        create_default_admin()
        
        # Clear old sessions to force re-login on startup
        clear_old_sessions()

    # Start background task workers
    start_task_workers()

    # Print database status
    if db_exists:
        print("[INFO] [OK] Using existing database")
    else:
        print("[INFO] [OK] New database created successfully")
    
    print("""
    ========================================================
    PLACEMENT MANAGEMENT SYSTEM - Flask Application
    ========================================================
    
    [*] Default Admin Credentials:
        Username: admin
        Password: admin
    
    [*] Starting server at http://localhost:5000
    
    [*] Features:
        - Admin Dashboard with full control
        - Teacher Dashboard with class management
        - Student CRUD operations
        - Automatic placement notifications
        - WhatsApp + Voice message automation
    
    [*] WhatsApp Bot Status:
        Using Twilio WhatsApp API
    """)
    
    # Initialize Twilio WhatsApp
    print("[STARTUP] [*] Initializing Twilio WhatsApp...")
    from whatsapp_automation import initialize_whatsapp
    if initialize_whatsapp():
        print("[STARTUP] [OK] Twilio WhatsApp integration ready!")
    else:
        print("[STARTUP] [!] Twilio credentials missing or invalid")
        print("[STARTUP] [!] Please check your .env file")
    
    print("""
    [*] Press Ctrl+C to stop the application
    """)
    
    app.run(debug=True, host='0.0.0.0', port=5000, use_reloader=False)
